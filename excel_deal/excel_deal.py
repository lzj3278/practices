# -*- coding:utf-8 -*-

from openpyxl import load_workbook
from collections import OrderedDict
from datetime


# from multiprocessing import Pool
# from functools import partial

def date_confirm():


def calculate_hours(wb, dics):
    sheets = wb.get_sheet_names()
    estimate_times = []
    for sheet in sheets:
        if sheet not in [u'目录', u'工时汇总']:
            ws = wb.get_sheet_by_name(sheet)
            location = dics['estimate_time'] + '15'
            time1 = ws[location].value
            if time1 is None:
                time1 = 0
            estimate_times.append(time1)
    print(sum(estimate_times))

if __name__ == '__main__':
    name_dics = [dict(estimate_time='D', actual_time='E'),
                 dict(estimate_time='F', actual_time='G'), ]
    wb = load_workbook(filename=u'2016.xlsx')
    for dics in name_dics:
        # print dics
        calculate_hours(wb, dics)
